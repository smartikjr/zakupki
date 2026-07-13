"""Выгрузка реестра НОСТРОЙ в .xlsx (один лист, отдельный от остальных приложений)."""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET = "НОСТРОЙ (реестр СРО)"

COLUMNS = [
    ("Дата загрузки", "loaded_at", 18),
    ("Компания", "company_name", 32),
    ("Компания (DaData)", "company", 30),
    ("ИНН", "inn", 15),
    ("ОГРН", "ogrn", 18),
    ("СРО", "sro_name", 30),
    ("Регион", "region", 22),
    ("Статус членства", "status_sro", 16),
    ("Дата вступления", "admission_date", 16),
    ("Руководитель", "director", 26),
    ("Статус (DaData)", "status", 12),
    ("Осн. ОКВЭД", "okved", 12),
    ("Адрес", "address", 40),
    ("Ссылка", "link", 45),
]


def load_existing(path: str) -> list[dict]:
    """Читает ранее сохранённые записи обратно в список словарей (не повторять между прогонами)."""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if SHEET not in wb.sheetnames:
        return []
    ws = wb[SHEET]
    key_by_title = {title: key for title, key, _w in COLUMNS}
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for title, val in zip(headers, row):
            key = key_by_title.get(title)
            if key:
                rec[key] = val if val is not None else ""
        rows.append(rec)
    return rows


def export(rows: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (title, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row_idx, row in enumerate(rows, start=2):
        row.setdefault("loaded_at", now)
        for col_idx, (_t, key, _w) in enumerate(COLUMNS, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "link" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows), 1) + 1}"
    wb.save(path)
