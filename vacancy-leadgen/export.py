"""Выгрузка результатов поиска подрядчиков в .xlsx (openpyxl)."""

from datetime import datetime
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COMPANY_COLUMNS = [
    ("Дата загрузки", "loaded_at", 18),
    ("Компания", "company", 35),
    ("ИНН", "inn", 15),
    ("Регион (DaData)", "region", 22),
    ("Руководитель", "director", 26),
    ("Статус", "status", 12),
    ("Осн. ОКВЭД", "okved", 12),
    ("Должности (вакансии)", "titles_str", 40),
    ("Кол-во вакансий", "vacancy_count", 14),
    ("Регион поиска (hh.ru)", "regions_str", 30),
    ("Последняя вакансия", "last_published", 18),
    ("Профиль на hh.ru", "hh_url", 40),
]

LINK_COLUMNS = [
    ("Регион", "region", 22),
    ("Категория", "category", 30),
    ("Запрос", "query", 45),
    ("Яндекс", "yandex_url", 45),
    ("Google", "google_url", 45),
]


def _header(ws, columns):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (title, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_search_links(categories: list[str], regions: list[str]) -> list[dict]:
    """Готовые кликабельные запросы для ручного поиска (п.9 методики) —
    сами результаты поиска не парсим (см. README: почему)."""
    out = []
    for region in regions:
        for cat in categories:
            query = f"{cat} {region}"
            out.append(
                {
                    "region": region,
                    "category": cat,
                    "query": query,
                    "yandex_url": f"https://yandex.ru/search/?text={quote_plus(query)}",
                    "google_url": f"https://www.google.com/search?q={quote_plus(query)}",
                }
            )
    return out


def export(companies: list[dict], links: list[dict], path: str) -> None:
    wb = Workbook()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws1 = wb.active
    ws1.title = "Компании (HH.ru)"
    _header(ws1, COMPANY_COLUMNS)
    for row_idx, c in enumerate(companies, start=2):
        c.setdefault("loaded_at", now)
        c["titles_str"] = ", ".join(sorted(c.get("titles", [])))
        c["regions_str"] = ", ".join(sorted(c.get("regions", [])))
        for col_idx, (_title, key, _w) in enumerate(COMPANY_COLUMNS, start=1):
            val = c.get(key, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "hh_url" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
    ws1.freeze_panes = "A2"
    if companies:
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(COMPANY_COLUMNS))}{len(companies) + 1}"

    ws2 = wb.create_sheet("Ссылки для поиска")
    _header(ws2, LINK_COLUMNS)
    for row_idx, l in enumerate(links, start=2):
        for col_idx, (_title, key, _w) in enumerate(LINK_COLUMNS, start=1):
            val = l.get(key, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in ("yandex_url", "google_url") and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
    ws2.freeze_panes = "A2"
    if links:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(LINK_COLUMNS))}{len(links) + 1}"

    wb.save(path)
