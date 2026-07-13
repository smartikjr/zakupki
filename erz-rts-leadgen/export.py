"""Выгрузка лидов ЕРЗ.РФ и РТС-тендер в один .xlsx с двумя отдельными листами."""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RTS_SHEET = "РТС-тендер"
ERZ_SHEET = "ЕРЗ.РФ (застройщики)"

RTS_COLUMNS = [
    ("Дата загрузки", "loaded_at", 18),
    ("Реестровый №", "reg_number", 22),
    ("Предмет закупки", "object", 55),
    ("Заказчик", "customer_name", 30),
    ("Компания (DaData)", "company", 30),
    ("ИНН", "inn", 15),
    ("Регион", "region", 22),
    ("Руководитель", "director", 26),
    ("Статус", "status", 12),
    ("Осн. ОКВЭД", "okved", 12),
    ("НМЦК", "price", 18),
    ("Сроки", "dates", 28),
    ("Адрес", "address", 40),
    ("Ссылка", "link", 45),
]

ERZ_COLUMNS = [
    ("Дата загрузки", "loaded_at", 18),
    ("Застройщик", "developer", 30),
    ("Компания (DaData)", "company", 30),
    ("ИНН", "inn", 15),
    ("Рейтинг РФ", "rank", 12),
    ("Регион (вид рейтинга)", "region", 18),
    ("Доля в регионе", "share_percent", 14),
    ("Строится в регионе", "volume_building", 18),
    ("Всего проектов в ЕРЗ", "total_projects", 18),
    ("Год основания", "founded_year", 14),
    ("Руководитель", "director", 26),
    ("Статус", "status", 12),
    ("Осн. ОКВЭД", "okved", 12),
    ("Адрес", "address", 40),
    ("Ссылка", "link", 45),
]


def _write_sheet(wb: Workbook, title: str, columns: list[tuple[str, str, int]], rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (col_title, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row_idx, row in enumerate(rows, start=2):
        row.setdefault("loaded_at", now)
        for col_idx, (_t, key, _w) in enumerate(columns, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "link" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(rows), 1) + 1}"


def load_existing(path: str, sheet: str, columns: list[tuple[str, str, int]]) -> list[dict]:
    """Читает ранее сохранённый лист обратно в список словарей (не повторять лиды между прогонами)."""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    key_by_title = {title: key for title, key, _w in columns}
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for title, val in zip(headers, row):
            key = key_by_title.get(title)
            if key:
                rec[key] = val if val is not None else ""
        rows.append(rec)
    return rows


def export(erz_rows: list[dict], rts_rows: list[dict], path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # убираем дефолтный пустой лист — оба листа создаём сами
    _write_sheet(wb, ERZ_SHEET, ERZ_COLUMNS, erz_rows)
    _write_sheet(wb, RTS_SHEET, RTS_COLUMNS, rts_rows)
    wb.save(path)
