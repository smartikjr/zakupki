"""Выгрузка списка лидов в .xlsx (openpyxl)."""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Дата загрузки", "loaded_at", 18),
    ("Закон", "law", 8),
    ("Реестровый №", "reg_number", 22),
    ("Предмет закупки", "object", 55),
    ("Заказчик (из ЕИС)", "customer_name", 30),
    ("Компания (DaData)", "company", 30),
    ("ИНН", "inn", 15),
    ("Регион", "region", 22),
    ("Руководитель", "director", 26),
    ("Статус", "status", 12),
    ("Осн. ОКВЭД", "okved", 12),
    ("НМЦК", "price", 18),
    ("Сроки", "dates", 28),
    ("Адрес", "address", 40),
    ("Ссылка ЕИС", "link", 45),
]


def export(leads: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (title, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for row_idx, lead in enumerate(leads, start=2):
        lead.setdefault("loaded_at", now)
        for col_idx, (_title, key, _w) in enumerate(COLUMNS, start=1):
            val = lead.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "link" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(leads) + 1}"
    wb.save(path)
